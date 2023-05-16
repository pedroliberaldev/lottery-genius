import openpyxl.utils.exceptions
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import exceptions

import os
import httpx
import random
import logging
import requests


def download_file():
    headers = {
        'authority': 'pagead2.googlesyndication.com',
        'accept': '*/*',
        'accept-language': 'pt-BR,pt;q=0.9',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'referer': 'https://googleads.g.doubleclick.net/',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Linux"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'no-cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari\
                      /537.36',
        'x-client-data': 'CIy2yQEIo7bJAQipncoBCIeQywEIk6HLAQiHoM0BCO2qzQE=',
    }

    response = requests.get(
        'https://pagead2.googlesyndication.com/pcs/activeview?xai=AKAOjssAbbL3BLsdFcIODXnOe2vbXCZoMtw-ZMZ39XLUt68YT1ugW\
        ZPFN0y1z6SGe21_c0aSDTGv7LRfWiQALrZyvUt7YVRUjwpKxlZyedx4D8Qk6_pGPSzmSjKlcKp4yCjz9JRP5IzMmA&sai=AMfl-YS3RF3sP-25u\
        C09_dOs6qFL81WMFHtoNrbVOHCpS4fnc4k1vUi5E2zfLxE2tCHBe-UbT5F8FEu1EMIkknZ6Afm2fz_2ghd4Lvnc1-xuqO_4Bqg458l4KB09GTEG\
        S93BJALhGQ&sig=Cg0ArKJSzBdbBorEHXtUEAE&id=lidar2&mcvt=1004&p=1,0,281.015625,1058.015625&mtos=0,0,978,1004,2782&\
        tos=0,0,978,26,3908&v=20230510&bin=7&avms=nio&bs=0,0&mc=0.56&if=1&vu=1&app=0&itpl=22&adk=281220856&rs=2&la=1&cr\
        =0&uach=WyJMaW51eCIsIjYuMi4xNSIsIng4NiIsIiIsIjExMy4wLjU2NzIuOTIiLFtdLDAsbnVsbCwiNjQiLFtbIkdvb2dsZSBDaHJvbWUiLCI\
        xMTMuMC41NjcyLjkyIl0sWyJDaHJvbWl1bSIsIjExMy4wLjU2NzIuOTIiXSxbIk5vdC1BLkJyYW5kIiwiMjQuMC4wLjAiXV0sMF0%3D&vs=4&r=\
        v&rst=1684167594463&rpt=2125&met=mue&wmsd=0&pbe=0&vae=0&spb=0',
        headers=headers,
    )


def download_file_out(url, output_path):
    headers = {
        'authority': 'pagead2.googlesyndication.com',
        'accept': '*/*',
        'accept-language': 'pt-BR,pt;q=0.9',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'referer': 'https://googleads.g.doubleclick.net/',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Linux"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'no-cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0Safari\
                      /537.36',
        'x-client-data': 'CIy2yQEIo7bJAQipncoBCIeQywEIk6HLAQiHoM0BCO2qzQE=',
    }

    with httpx.Client() as client:
        response = client.get(
            'https://pagead2.googlesyndication.com/pcs/activeview?xai=AKAOjssAbbL3BLsdFcIODXnOe2vbXCZoMtw-ZMZ39XLUt68YT\
            1ugWZPFN0y1z6SGe21_c0aSDTGv7LRfWiQALrZyvUt7YVRUjwpKxlZyedx4D8Qk6_pGPSzmSjKlcKp4yCjz9JRP5IzMmA&sai=AMfl-YS3R\
            F3sP-25uC09_dOs6qFL81WMFHtoNrbVOHCpS4fnc4k1vUi5E2zfLxE2tCHBe-UbT5F8FEu1EMIkknZ6Afm2fz_2ghd4Lvnc1-xuqO_4Bqg4\
            58l4KB09GTEGS93BJALhGQ&sig=Cg0ArKJSzBdbBorEHXtUEAE&id=lidar2&mcvt=1004&p=1,0,281.015625,1058.015625&mtos=0,\
            0,978,1004,2782&tos=0,0,978,26,3908&v=20230510&bin=7&avms=nio&bs=0,0&mc=0.56&if=1&vu=1&app=0&itpl=22&adk=28\
            1220856&rs=2&la=1&cr=0&uach=WyJMaW51eCIsIjYuMi4xNSIsIng4NiIsIiIsIjExMy4wLjU2NzIuOTIiLFtdLDAsbnVsbCwiNjQiLFt\
            bIkdvb2dsZSBDaHJvbWUiLCIxMTMuMC41NjcyLjkyIl0sWyJDaHJvbWl1bSIsIjExMy4wLjU2NzIuOTIiXSxbIk5vdC1BLkJyYW 5kIiwiM\
            jQuMC4wLjAiXV0sMF0%3D&vs=4&r=v&rst=1684167594463&rpt=2125&met=mue&wmsd=0&pbe=0&vae=0&spb=0',
            headers=headers,
        )

        if response.status_code == 200:
            with open(output_path, "wb") as file:
                file.write(response.content)
            print("Download concluído.")
        else:
            print("Falha ao fazer o download.")


# Function to read the xlsx with lottery results history
def read_file() -> object:
    try:
        # Get the current path to set the filename and create a new workbook to importing data
        filename: str = os.path.join(os.getcwd(), "results_file", "results.xlsx")
        file_workbook = load_workbook(filename)

        # Return the generated workbook
        return file_workbook
    # Exception in case of can not read the file
    except exceptions.InvalidFileException:
        logging.error("Can not read the workbook file")
        return False


# Function to create the recurrence xlsx file. This file will be used to select best numbers
def create_data(file_workbook: object) -> object:
    # Create the new workbook, select and rename the active sheet
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Data"

    # Define the base structure of new workbook (B1 and all A column)
    for interactor in range(1, 62):
        current_a_collumn = "A" + str(interactor)
        current_b_collumn = "B" + str(interactor)

        if interactor == 1:
            new_sheet[current_a_collumn] = "Numbers"
            new_sheet[current_b_collumn] = "Recurrence"
        else:
            new_sheet[current_a_collumn] = interactor - 1

    # Access file workbook sheet and checks the total number of draws made
    file_sheet = file_workbook["Sheet1"]
    draws_total = file_sheet['A2'].value

    # Navigate file workbook and count the recurrence of all numbers
    for row in file_sheet.iter_rows(min_row=2, min_col=3, max_col=8, max_row=draws_total, values_only=True):
        for cell in row:
            target_cell = "B" + str(cell + 1)
            check_cell_value = new_sheet[target_cell].value

            if (str(check_cell_value) == 'None') or (str(check_cell_value) == 'none'):
                new_sheet[target_cell] = 1
            else:
                new_sheet[target_cell] = int(new_sheet[target_cell].value) + 1

    # Save the new workbook
    filename: str = os.path.join(os.getcwd(), "results_file", "recurrence.xlsx")
    new_workbook.save(filename)

    # Return the recurrence workbook
    return new_workbook


# Function to create the sorted array with lottery history numbers
def create_data_sorted(file_workbook: object) -> object:
    got_numbers = []

    try:
        # Access file workbook sheet and checks the total number of draws made
        file_sheet = file_workbook["Data"]

        # Get all workbook numbers
        for row in file_sheet.iter_rows(min_row=2, min_col=1, max_row=61, values_only=True):
            for cell in row:
                target_cell = "B" + str(cell + 1)
                check_cell_value = file_sheet[target_cell].value

                if str(check_cell_value) == 'None' or str(check_cell_value) == 'none':
                    continue
                else:
                    got_numbers.append(check_cell_value)

        # Order original array and get indexes
        sorted_arrays = sorted(range(len(got_numbers)), key=lambda i: got_numbers[i])

        # Inverter indexes
        inverted_sorted_arrays = sorted_arrays[::-1]

        # Create the final array
        sorted_numbers = [i + 1 for i in inverted_sorted_arrays]

        # Return the final and sorted array
        return sorted_numbers
    except exceptions.InvalidFileException:
        logging.error("Can not create sorted array")
        return False


# Function to create the lottery game based on users choices
def create_game(sorted_numbers, total_of_balls, total_of_best_numbers, qtd_high_recurrence_numbers):
    # The final game numbers array
    my_game = []

    # Low and high limits os lottery numbers
    low_limit = 1
    high_limit = 60

    # Counters to the numbers of high and low recurrence numbers
    high_recurrence_numbers = 0
    low_recurrence_numbers = 0

    # Define how many low recurrence balls should be randomized
    qtd_low_recurrence_numbers = total_of_balls - qtd_high_recurrence_numbers

    while len(my_game) < total_of_balls:
        random_number = random.randint(low_limit, high_limit)

        if random_number not in my_game:
            if random_number in sorted_numbers[:total_of_best_numbers]:
                if high_recurrence_numbers < qtd_high_recurrence_numbers:
                    my_game.append(random_number)
                    high_recurrence_numbers += 1
            else:
                if low_recurrence_numbers < qtd_low_recurrence_numbers:
                    my_game.append(random_number)
                    low_recurrence_numbers += 1

    # Sort the game numbers
    my_game = sorted(my_game)

    # Return the final game numbers
    return my_game


url = "https://redeloteria.com.br/mega-sena/todos-os-resultados-da-mega-sena/29275"
download_file_out(url, output_path=os.path.join(os.getcwd(), "results_file", "downloaded_results.xlsx"))

returned_workbook = read_file()
recurrence_workbook = create_data(returned_workbook)
recurrence_sorted_numbers = create_data_sorted(recurrence_workbook)
final_game = create_game(recurrence_sorted_numbers, 6, 20, 4)
