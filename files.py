from openpyxl import load_workbook
from openpyxl import Workbook
import os
import httpx


def download_file(url, output_path):
    with httpx.Client() as client:
        response = client.get(url)

        if response.status_code == 200:
            with open(output_path, "wb") as file:
                file.write(response.content)
            print("Download concluído.")
        else:
            print("Falha ao fazer o download.")


def read_file() -> object:
    """

    :rtype: object
    """
    # Get the current path to set the filename and create a new workbook to importing data
    filename: str = os.path.join(os.getcwd(), "results_file", "results.xlsx")
    file_workbook = load_workbook(filename)

    return file_workbook


def create_data(file_workbook) -> object:
    # Create the new workbook, select and rename the active sheet
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Data"

    for interactor in range(1, 62):
        current_a_collumn = "A" + str(interactor)
        current_b_collumn = "B" + str(interactor)

        if (interactor == 1):
            new_sheet[current_a_collumn] = "Numbers"
            new_sheet[current_b_collumn] = "Recurrence"
        else:
            new_sheet[current_a_collumn] = interactor - 1

    # Access file workbook sheet and checks the total number of draws made
    file_sheet = file_workbook["Sheet1"]
    draws_total = file_sheet['A2'].value

    for row in file_sheet.iter_rows(min_row=2, min_col=3, max_col=8, max_row=draws_total, values_only=True):
        for cell in row:
            target_cell = "B" + str(cell + 1)
            check_cell_value = new_sheet[target_cell].value

            if (str(check_cell_value) == 'None') or (str(check_cell_value) == 'none'):
                new_sheet[target_cell] = 1
            else:
                new_sheet[target_cell] = int(new_sheet[target_cell].value) + 1

    # Save the new workbook
    filename: str = os.path.join(os.getcwd(), "results_file", "recurrence.xlsx")
    new_workbook.save(filename)

    return new_workbook
